"""
- Helper functions for data export
"""

import pandas as pd
import numpy as np
import openpyxl, math
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Color, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt

# Set Styles
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
no_border = Border(left=Side(style=None),
                   right=Side(style=None),
                   top=Side(style=None),
                   bottom=Side(style=None))
fontStyle = Font(name='Calibri', size='10')
zoom_level = 85


def create_empty_excel_template(out_fname, worksheet_list):
    """
    Parameters
    ----------
        out_fname: str, output excel filename, with location
        worksheet_list: list of worksheets to be created within the excel sheet
    Returns
    -------
        creates an empty excel file with the worksheets mentioned
    """
    # Create a Workbook
    wb = Workbook()
    
    # Remove the default sheet and Create New sheets
    wb.remove(wb.active)
    for sheet in worksheet_list:
        ws_sheet = wb.create_sheet(title=sheet)
        
    # Save the Workbook
    wb.save(filename=out_fname)
    
    
def export_data_to_excel(data, out_fname, sheetName, pct_col_list=None):
    """
    Parameters
    ----------
        data: pandas dataframe
        out_fname: str, location and name of output file, should be pre-existing
        sheetName: str, name of sheet in excel, should be pre-existing
        pct_col_list: optional parameter; str or list of columns that needs % formatting
    Returns
    -------
        exports the dataframe to the sheetName within out_fname, with the pct_col_list with % formatting
    """
    # Open Workbook and Load Worksheet
    wb = openpyxl.load_workbook(out_fname)
    ws = wb[sheetName]
    
    # Write Data to Worksheet
    data = data.astype(object)
    rows = dataframe_to_rows(data, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        
    # Percentage Formatting
    if pct_col_list != None:
        if type(pct_col_list) == str:
            pct_col_list = list(pct_col_list.split())
            
        # Change the Formatting
        for pct_col in pct_col_list:
            i=0
            for cell in ws[1]:
                i += 1
                if cell.value == pct_col:
                    break
                    
            for cell in ws[get_column_letter(i)]:
                cell.number_format = FORMAT_PERCENTAGE
                
    # Set Width of Columns
    for column_cells in ws.columns:
        length = max(len(_as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length+4
        
    # Make Header Bold and Centrally Aligned
    for _cell in ws[1]:
        _cell.font = Font(bold=True)
        _cell.alignment = Alignment(horizontal='center')
        _cell.fill = PatternFill(patternType='solid',
                                 fill_type='solid',
                                 fgColor=Color('00C0C0C0'))
        
    # Set Borderlines
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            
    # Other Adjustments
    ws.sheet_view.zoomScale = zoom_level
    ws.sheet_view.showGridLines = False
    
    # Save Worksheet
    wb.save(out_fname)
    
    
def export_coarse_classing_data(class_df, outfile, sheetName):
    """
    Parameters
    ----------
        class_df: pandas dataframe, coarse classing data
        outfile: str, location and name of the output excel file
        sheetName: str, name of worksheet
    Returns
    -------
        exports the coarse classing data into the mentioned excel file worksheet
    """
    # Data Export
    writer = pd.ExcelWriter(outfile, engine='xlsxwriter')
    line_counter = 2
    var_list = list(class_df['VAR_NAME'].unique())
    for var in var_list:
        _class_df = class_df[class_df['VAR_NAME'] == var]
        _class_df.to_excel(writer, sheet_name=sheetName, startrow=line_counter, startcol=1, header=True, index=False)
        line_counter += len(_class_df) + 3
    writer.save()
    
    # Format Excel Sheet
    wb = openpyxl.load_workbook(outfile)
    ws = wb[sheetName]
    
    # % Formatting
    pct_col_list=['ROWP_TOT', 'PER_RESP', 'PER_NON_RESP', 'RESP_RATE']
    for pct_col in pct_col_list:
        i=0
        for cell in ws[3]:
            i += 1
            if cell.value == pct_col:
                break
        for cell in ws[get_column_letter(i)]:
            if cell.value != pct_col:
                cell.number_format = FORMAT_PERCENTAGE
                
    ws.sheet_view.zoomScale = zoom_level
    wb.save(outfile)

    
def plot_woe_graphs(var, woe_df, out_loc):
    """
    Parameters
    ----------
        var: str, variable name for which WOE needs to be plotted
        woe_df: pandas dataframe, with WOE values
        out_loc: str, location where graphs will be saved in png format
    Returns
    -------
        creates and saves WOE and response rate graphs for the column mentioned
    """
    # Create Data Series
    woe_df['Category1'] = woe_df['VAR_BINS'].apply(lambda x: str(x))
    xpoints = woe_df['Category1']
    y1points = woe_df['LN_ODDS']
    y2points = woe_df['RESP_RATE']
    
    # Plot Data
    fig, ax1 = plt.subplots()
    ax2 = ax1.twinx()
    ax1.plot(xpoints, y1points, '-cD', label='WOE')
    ax2.plot(xpoints, y2points, '-rD', label='WOE')
    
    # Format Plot
    plt.title(f'WOE & REsponse Rate: {var}')
    ax1.set_xlabel(f'{var} Buckets')
    ax1.set_ylabel('Weight of Evidence')
    ax2.set_ylabel('Response Rate')
    fig.legend(loc=4)
    ax1.tick_params(axis='x', labelrotation=45)
    plt.tight_layout()
    
    # Show/Save Plots
#     plt.save(f'{out_loc}/graph/woe_buckets_{var}.png', dpi=70)
    plt.savefig(f'{out_loc}/graph/woe_buckets_{var}.png', dpi=70)
    plt.close(fig)
    
    
def export_fine_classing_data(f_class_df, out_loc, out_fname):
    """
    Parameters
    ----------
        class_df: pandas dataframe, coarse classing data
        outfile: str, location and name of the output excel file
        sheetName: str, name of worksheet
    Returns
    -------
        exports the coarse classing data into the mentioned excel file worksheet
    """
    # Data Export
    writer = pd.ExcelWriter(f'{out_loc}/{out_fname}', engine='xlsxwriter')
    line_counter = 2
    var_list = list(f_class_df['VAR_NAME'].unique())
    for var in var_list:
        _f_class_df = f_class_df[f_class_df['VAR_NAME'] == var]
        _f_class_df.to_excel(writer, sheet_name='Fine Classing Data', startrow=line_counter, startcol=1, header=True, index=False)
        if (len(_f_class_df) > 17):
            line_counter += len(_f_class_df) + 2
        else:
            line_counter = line_counter + 20
    writer.save()
    
    # Graph Export
    line_counter = 2
    for var in var_list:
        _f_class_df = f_class_df[f_class_df['VAR_NAME'] == var]
        
        # Save Graph
        plot_woe_graphs(var, _f_class_df, out_loc)
        img = openpyxl.drawing.image.Image(f'{out_loc}/graph/woe_buckets_{var}.png')
        
        wb = openpyxl.load_workbook(f'{out_loc}/{out_fname}')
        ws = wb.active
        
        img.anchor = 'P'+str(line_counter)
        ws.add_image(img)
        ws.sheet_view.zoomScale = zoom_level
        wb.save(f'{out_loc}/{out_fname}')
        wb.close()
        
        # Update Buffer Space
        if (len(_f_class_df) > 17):
            line_counter += len(_f_class_df) + 2
        else:
            line_counter = line_counter + 20
        
        
    
def _as_text(value):
    if value is None:
        return('')
    else:
        return(str(value))